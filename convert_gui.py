import sys, csv, chardet, os
from tkinter import *
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook


root = Tk()
input_file_name = None
xlsx_name = None
folder_name = None

def Convert():

    global input_file_name, xlsx_name, folder_name
    
    csv_name = input_file_name.get()
    out_folder_name = folder_name.get()
    out_xlsx_name = xlsx_name.get()
    csv_file_name = csv_name.split('.')
    
    try:
        with open(csv_name, 'rb') as file:
            result = chardet.detect(file.read())
            with open(csv_name, 'r', encoding=result['encoding']) as file:
                csv_data = []
                reader = csv.reader(file)
                for row in reader:
                    csv_data.append(row)

    except FileNotFoundError:
        messagebox.showerror('File Not Found', f'{os.path.basename(csv_name)} が見つかりませんでした。')
        return

    file_name = ''

    if out_xlsx_name == '':
        for i, name in enumerate(csv_file_name):
            if i != len(csv_file_name) - 2:
                file_name += name + '.'
            else:
                file_name += name
                break
    else:
        if out_xlsx_name.split('.')[-1].lower() != 'xlsx':
            for i, name in enumerate(out_xlsx_name.split('.')):
                if i != len(out_xlsx_name.split('.')) - 1:
                    file_name += name + '.'
                else:
                    file_name += name
                    break
        else:
            for i, name in enumerate(out_xlsx_name.split('.')):
                if i != len(out_xlsx_name.split('.')) - 2:
                    file_name += name + '.'
                else:
                    file_name += name
                    break

    try:
        out_path = os.path.join(out_folder_name, f'{os.path.basename(file_name)}.xlsx')
        with open(out_path, 'r'):
            ret = messagebox.askyesno('上書き確認', f'{os.path.basename(file_name)}.xlsx が既に存在します。\n上書きしてよろしいですか？')
            if ret == False:
                messagebox.showinfo('キャンセル', '上書きをキャンセルしました')
                return
    except FileNotFoundError:
        pass

    wb = Workbook()

    try:
        wb.save(out_path)
    except PermissionError:
        messagebox.showerror('ファイル使用中', f'{os.path.basename(file_name)}.xlsx が開かれているため上書きできません。\nExcelファイルを閉じるか、出力するファイル名を変更してください。')
        sys.exit(1)

    wb = load_workbook(out_path)

    ws = wb['Sheet']

    for i, row in enumerate(csv_data):
        for j, column in enumerate(row):
            ws.cell(row=i+1, column=j+1).value = column

    wb.save(out_path)
    messagebox.showinfo('保存完了', f'{file_name}.xlsx を保存しました')


def ChooseCSVFile():
    
    global input_file_name, folder_name
    
    path = os.getcwd()
    dir_name = os.path.dirname(path)
    
    type_of_file = [('CSVファイル', '*.csv')]
    
    filename = filedialog.askopenfilename(filetypes=type_of_file, initialdir=dir_name)
    
    if filename:
        input_file_name.delete(0, END)
        input_file_name.insert(0, filename)
    
    if folder_name.get() == '':
        folder_name.delete(0, END)
        folder_name.insert(0, os.path.dirname(filename))
    
    if filename != '':
        Button(text="xlsxファイルへ変換", command=Convert).grid(row=3, column=0, padx=10, pady=10, sticky="w", columnspan=2)


def ChooseFolder():
    
    global input_file_name, folder_name
    
    path = os.getcwd()
    dir_name = os.path.dirname(path)
    
    foldername = filedialog.askdirectory(initialdir=dir_name)
    
    if foldername:
        folder_name.delete(0, END)
        folder_name.insert(0, foldername)


def WindowMain():
    
    global input_file_name, xlsx_name, folder_name
    
    root.title("CSV to Excel Converter")
    root.geometry("600x400")
    
    Label(text="入力CSVファイル:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
    input_file_name = Entry(width=30, justify="left")
    input_file_name.grid(row=0, column=1, pady=10)
    Button(text="ファイル選択", command=ChooseCSVFile).grid(row=0, column=2, pady=10)
    Label(text="出力フォルダ:").grid(row=1, column=0, padx=10, sticky="w")
    folder_name = Entry(width=30, justify="left")
    folder_name.grid(row=1, column=1, pady=10)
    Button(text="フォルダ選択", command=ChooseFolder).grid(row=1, column=2, pady=10)
    Label(text="出力ファイル名:").grid(row=2, column=0, padx=10, sticky="w")
    xlsx_name = Entry(width=30, justify="left")
    xlsx_name.grid(row=2, column=1, pady=10)
    
    root.mainloop()

if __name__ == '__main__':
    WindowMain()
