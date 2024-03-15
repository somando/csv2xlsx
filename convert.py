import sys, csv, chardet
from openpyxl import Workbook, load_workbook

args = sys.argv

if len(args) < 2 or len(args) > 3 or (len(args) == 2 and (args[1] == '-h' or args[1] == '--help' or args[1].split('.')[-1].lower() != 'csv')) or (len(args) == 3 and (args[1].split('.')[-1].lower() != 'csv' or args[2].split('.')[-1].lower() != 'xlsx')):
    print('使い方: python convert.py <csvファイル名> <保存ファイル名 *任意>')
    sys.exit(1)

csv_file_name = args[1].split('.')

try:
    with open(args[1], 'rb') as file:
        result = chardet.detect(file.read())
        with open(args[1], 'r', encoding=result['encoding']) as file:
            csv_data = []
            reader = csv.reader(file)
            for row in reader:
                csv_data.append(row)
    
    print(f'{args[1]} を読み込みました')

except FileNotFoundError:
    print(f'{args[1]} が見つかりません')
    sys.exit(1)

file_name = ''

if len(args) == 2:
    for i, name in enumerate(csv_file_name):
        if i != len(csv_file_name) - 2:
            file_name += name + '.'
        else:
            file_name += name
            break
else:
    for i, name in enumerate(args[2].split('.')):
        if i != len(args[2].split('.')) - 2:
            file_name += name + '.'
        else:
            file_name += name
            break

try:
    with open(f'{file_name}.xlsx', 'r'):
        check = input(f'{file_name}.xlsx が既に存在します。\n上書きしてよろしいですか？(y/N) ... ')
        if check.lower() == 'n' or check.lower() == 'no' or check == '':
            sys.exit(1)
except FileNotFoundError:
    pass

wb = Workbook()

try:
    wb.save(f'{file_name}.xlsx')
except PermissionError:
    print(f'{file_name}.xlsx が開かれているため上書きできません。Excelファイルを閉じるか、出力するファイル名を変更してください。')
    sys.exit(1)

wb = load_workbook(f'{file_name}.xlsx')

wb.create_sheet(title=f'{file_name}', index=0)
wb.remove(wb['Sheet'])

ws = wb[f'{file_name}']

for i, row in enumerate(csv_data):
    for j, column in enumerate(row):
        ws.cell(row=i+1, column=j+1).value = column

wb.save(f'{file_name}.xlsx')
print(f'{file_name}.xlsx を保存しました')
